from matplotlib import pyplot as plt
from openpyxl import load_workbook, Workbook


def load_data(data_name):
    """
    Загружает xlsx файлы с данными трёх датчиков (T1, T2, T3) для дальнейшей обработки.

    :param data_name: Название серии измерений (н-р, 3-В4_deformation_2018_6MHz_sensor_3).
    :type data_name: str

    :return: 3 объекта типа Workbook, соответствующие датчикам Τ1, Τ2, Τ3.
    :rtype: tuple[Workbook, Workbook, Workbook]
    """

    wb_1 = load_workbook(f'./measurements/{data_name}/{data_name} (Т1).xlsx', data_only=True, read_only=True)
    wb_2 = load_workbook(f'./measurements/{data_name}/{data_name} (Т2).xlsx', data_only=True, read_only=True)
    wb_3 = load_workbook(f'./measurements/{data_name}/{data_name} (Т3).xlsx', data_only=True, read_only=True)

    return wb_1, wb_2, wb_3


def get_rows_count(book, sheet_name):
    """
    Возвращает количество строк в листе.

    :param book: Workbook, в котором находится нужный лист
    :type book: Workbook

    :param sheet_name: Название листа
    :type sheet_name: str

    :rtype: int
    """

    sheet = book[sheet_name]

    return len(list(sheet.rows))


def str_to_float_if_needed(val):
    """
    Иногда Excel выдаёт данные в формате str вместо float. Функция решает эту проблему.\n
    Также функция округляет значения до 4-х знаков после запятой.

    Пример:
    ' 18,9452 ' -> 18.9452

    :param val: Входные данные
    :type val: str | float

    :rtype: float
    """

    if isinstance(val, str):
        val = float(val[1:-1].replace(',', '.'))

    return round(val, 4)


def get_values(book, sheet_name):
    """
    Выгружает пару данных с листа в 2 массива, игнорируя ошибочные ячейки.\n
    При считывании используется функция str_to_float_if_needed.\n
    Считанные значения округляются до 4-х знаков после запятой.

    :param book: Workbook, в котором находятся данные.
    :type book: Workbook

    :param sheet_name: Название листа, с которого считываются данные.
    :type sheet_name: str

    :rtype: tuple[list[float | str], list[float]]
    """

    rows_amount = get_rows_count(book, sheet_name)
    sheet = book[sheet_name]
    val_1 = []
    val_2 = []

    for r in range(2, rows_amount+1):
        if sheet.cell(row=r, column=2).value not in ['ERROR!', 'ERROR! ']:
            val_1.append(str_to_float_if_needed(sheet.cell(row=r, column=2).value))
            val_2.append(round(sheet.cell(row=r, column=3).value, 4))

    return val_1, val_2


def mean_value(vals):
    """
    Находит среднее значение серии. Результат округляется до 4-х знаков после запятой.

    :param vals: Массив с данными формата float.
    :type vals: list[float]
    :rtype: float
    """

    mean_val = round(sum(vals) / len(vals), 4)
    return mean_val


def mean_abs_value(vals):
    """
        Находит среднее по модулю значение серии. Результат округляется до 4-х знаков после запятой.

        :param vals: Массив с данными формата float.
        :type vals: list[float]
        :rtype: float
        """

    vals = list(map(lambda x: abs(x), vals))
    mean_abs_val = round(sum(vals) / len(vals), 4)
    return mean_abs_val


def error_finder(vals_1, vals_2):
    """
    Определяет константу сдвига значений. Результат округляется до 2-х знаков после запятой.\n
    Выделяет пары с большой ошибкой, для оставшихся значений находит среднее - база, от которой находится сдвиг, находит средний сдвиг выделенных измерений.

    :param vals_1: Массив значений 1
    :type vals_1: list[float]

    :param vals_2: Массив значений 2
    :type vals_2: list[float]

    :rtype: float

    .. note::

        Если абсолютный сдвиг значения оказывается больше 0.35, то это значение игнорируется.
    """

    error_indexs = []
    delta_vals = []

    for l in range(len(vals_1)):
        delta_val = vals_1[l] - vals_2[l] * 2
        delta_vals.append(delta_val)

        if abs(delta_val) > 0.1:
            error_indexs.append(l)

    if not error_indexs:
        return 0.0
    else:
        delta_vals = []

        for l in range(len(vals_1)):
            if l not in error_indexs:
                delta_val = vals_1[l] - vals_2[l] * 2
                delta_vals.append(delta_val)

        mean_abs_delta_val = mean_abs_value(delta_vals)
        errors = []

        for v in error_indexs:
            new_val_1, new_val_2 = error_correction(vals_1[v], vals_2[v], mean_abs_delta_val)
            delta_new_val = new_val_1 - new_val_2 * 2

            if 0.1 <= abs(delta_new_val) <= 0.3:
                errors.append(delta_new_val)

        if not errors:
            return 10.0
        else:
            error = round(mean_abs_value(errors), 3)

        return error


def error_correction(val_1, val_2, err):
    """
    Корректирует значения на заданную константу по следующим правилам:\n
    Если ошибка между .1 и .4 -> значение 1 - константа;\n
    Если ошибка больше .4 -> значение 2 + константа;\n
    Если ошибка между -.4 и -.1 -> значение 1, значение 2 - константа;\n
    Если ошибка меньше -.4 -> значение 2 - константа.

    :param val_1: Значение 1
    :type val_1: float

    :param val_2: Значение 2
    :type val_2: float

    :param err: Константа сдвига
    :type err: float

    :return: Пару значений в порядке Значение 1, Значение 2
    :rtype: tuple[float, float]

    .. note::

        Ошибка всё ещё может остаться большой, в таких случаях можно повторно использовать корректировку или игнорировать такие значения.
    """

    delta_val = round(val_1 - val_2 * 2, 4)

    if abs(delta_val) >= 0.1:
        if 0.1 <= delta_val <= 0.4:
            val_1 -= err
        elif -0.4 <= delta_val <= -0.1:
            val_1 -= err
            val_2 -= err
        elif delta_val > 0.4:
            val_2 += err
        elif delta_val < -0.4:
            val_2 -= err

    return round(val_1, 4), round(val_2, 4)


file = input('Enter measurement name: ')
db = load_data(file)

wb = Workbook()

for i in range(3):
    n = 0
    wb['Sheet'].cell(row=((i * 4) + 2), column=1, value=f'T{(i + 1)}')
    wb['Sheet'].cell(row=((i * 4) + 2), column=2, value='dt')
    wb['Sheet'].cell(row=((i * 4) + 3), column=2, value='mt')
    wb['Sheet'].cell(row=((i * 4) + 4), column=2, value='err')
    wb['Sheet'].cell(row=((i * 4) + 5), column=2, value='UV|TV')

    shts = db[i].sheetnames

    for sht in shts:
        tn, t2 = get_values(db[i], sht)
        total_amount_of_values = len(tn)

        if total_amount_of_values >= 3:
            ta = list(map(lambda x: x * 2, t2))
            shift_constant = error_finder(tn, t2)
            delta_t = []

            new_tn = [0.0] * total_amount_of_values
            new_t2 = [0.0] * total_amount_of_values
            new_ta = [0.0] * total_amount_of_values

            for j in range(total_amount_of_values):
                new_tn[j], new_t2[j] = error_correction(tn[j], t2[j], shift_constant)
                new_ta[j] = new_t2[j] * 2
                delta_t.append(round(new_tn[j] - new_ta[j], 4))

            while max(list(map(lambda x: abs(x), delta_t))) > 0.1:
                for j in range(len(delta_t)):
                    if abs(delta_t[j]) > 0.1:
                        new_tn.pop(j)
                        new_t2.pop(j)
                        new_ta.pop(j)
                        delta_t.pop(j)
                        break

            used_amount_of_values = len(new_tn)
            mt = list(map(lambda x: round(x / 2, 4), new_tn))
            mean_delta_t = mean_abs_value(delta_t)
            mean_mt = mean_value(mt)
        else:
            mean_delta_t = 0.0
            mean_mt = 0.0
            shift_constant = 0.0
            used_amount_of_values = 0

        n += 1
        wb['Sheet'].cell(row=1, column=(n + 2), value=sht)
        wb['Sheet'].cell(row=((i * 4) + 2), column=(n + 2), value=mean_delta_t)
        wb['Sheet'].cell(row=((i * 4) + 3), column=(n + 2), value=mean_mt)
        wb['Sheet'].cell(row=((i * 4) + 4), column=(n + 2), value=shift_constant)
        wb['Sheet'].cell(row=((i * 4) + 5), column=(n + 2), value=f'{used_amount_of_values}|{total_amount_of_values}')

wb.save(f'./results/{file}/{file}.xlsx')
